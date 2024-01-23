# By Shmouel Illouz : Being lazy is knowing how to write scripts.

import openpyxl
import xml.etree.ElementTree as ET
import pandas as pd
import os

def find_sheet_with_most_rows(workbook):
    sheets = workbook.sheetnames
    sheet_with_most_rows = None
    max_rows = 0

    for sheet_name in sheets:
        sheet = workbook[sheet_name]
        num_rows = sheet.max_row

        if num_rows > max_rows:
            max_rows = num_rows
            sheet_with_most_rows = sheet_name

    return sheet_with_most_rows

def add_missing_elements(workbook, selected_sheet, column_A, list_to_compare, list_of_lists):
    sheet = workbook[selected_sheet]

    # Check for missing elements
    missing_elements = [element for element in list_to_compare if element not in column_A]

    if missing_elements:
        # Add missing elements to the end of column A
        last_row = sheet.max_row + 1
        for element in missing_elements:
            for details in list_of_lists:
                if details[0] == element:
                    sheet.cell(row=last_row, column=1, value=details[0])
                    sheet.cell(row=last_row, column=2, value=details[1])
                    sheet.cell(row=last_row, column=3, value=details[2])
                    sheet.cell(row=last_row, column=4, value="added")
                    last_row += 1

        print(f"The elements {missing_elements} have been added to the end of column A.")
    else:
        print("No missing elements found.")

# ******************************* MAIN ***********************************

# Ask the user for the Excel file path
excel_file_path = input("Please enter the path of the Excel file: ")

# Load the Excel file
workbook = openpyxl.load_workbook(excel_file_path)

# Find the sheet with the most rows
selected_sheet = find_sheet_with_most_rows(workbook)

if selected_sheet:
    sheet = workbook[selected_sheet]

    # Retrieve all values from column A into a list
    column_A = [cell.value for cell in sheet['A']]

    # List to compare
    xml_file_path = input("Enter the path of the XML file: ")
    # Parsing XML
    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    members_to_remove = ["et4ae5", "due__", "Google", "OSF_", "twilio", "odigo", "Didenjoy", "IndividualEmailResult", 
                       "ServiceTerritory", "ServiceResource", "SocialPost", "ServiceAppointment", "ResourceAbsence"
                       ,"WorkType", "SiqUserBlacklist", "SocialPost", "StreamActivity", "TableauHost", "UserEmailPreferred",
                       "VideoCall", "VoiceCall", "Waitlist"]
    members_list = []
    final_grand_list = []
    final_grand_list_members = []
    status_members_list = []

    df = pd.DataFrame()

    # Each type is defined by a name
    for types in root:
        name_type_list = ''
        members_list = []
        for elem in types:
            flag = 0
            elem_name = elem.tag.split('}')[-1]
            if elem_name == 'members':
                members_list.append(elem.text)
                for undesirables in members_to_remove:
                    if undesirables in elem.text:
                        flag = 1 
                status_members_list.append("Expected OK" if flag == 0 else "Not supported in SFOA")
            if elem_name == 'name':
                name_type_list = elem.text

        if name_type_list == 'EmailTemplate':
            status_members_list[-len(members_list):] = ["Data Scope"] * len(members_list)

        if name_type_list == 'Report':
            status_members_list[-len(members_list):] = ["Scope To Define"] * len(members_list)

        if name_type_list == 'Dashboard':
            status_members_list[-len(members_list):] = ["Scope To Define"] * len(members_list)

        list_of_lists = []

        # Building final_grand_list after status_members_list has been built
        for member in members_list:
            final_grand_list.append(member)
            final_grand_list_members.append(name_type_list)

    for i in range(len(final_grand_list)):
        details_member = [final_grand_list[i], final_grand_list_members[i], status_members_list[i]]
        # Adding the sublist to the main list
        list_of_lists.append(details_member)

    # Creating a Pandas DataFrame
    print("final_grand_list: " + str(len(final_grand_list)))
    print("final_grand_list_members: " + str(len(final_grand_list_members)))
    print("status_members_list: " + str(len(status_members_list)))
    print("list_of_lists: " + str(len(list_of_lists)))

    # df = pd.DataFrame({"Members": final_grand_list, "Type": final_grand_list_members  , "SFOA Status": status_members_list}) 

    list_to_compare = final_grand_list  # Replace this with your own list

    # Compare the lists and add missing elements
    add_missing_elements(workbook, selected_sheet, column_A, list_to_compare, list_of_lists)

    # Save the modifications to the Excel file
    workbook.save(excel_file_path)
    print("Modifications have been saved.")
else:
    print("No sheet found.")

# Close the Excel file
workbook.close()
