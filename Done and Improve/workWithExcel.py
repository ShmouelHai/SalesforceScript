import openpyxl

def find_sheet_with_most_rows(workbook):
    sheets = workbook.sheetnames
    sheet_with_most_rows = None
    max_rows = 0

    for sheet_name in sheets:
        sheet = workbook[sheet_name]
        num_rows = sheet.max_row

        if num_rows > max_rows:
            max_rows = num_rows
            sheet_with_most_rows = sheet_name

    return sheet_with_most_rows

def ajouter_elements_manquants(workbook, selected_sheet, column_A, list_to_compare):
    sheet = workbook[selected_sheet]

    # Vérifier les éléments manquants
    elements_manquants = [element for element in list_to_compare if element not in column_A]

    if elements_manquants:
        # Ajouter les éléments manquants à la fin de la colonne A
        last_row = sheet.max_row + 1
        for element in elements_manquants:
            sheet.cell(row=last_row, column=1, value=element)
            sheet.cell(row=last_row, column=2, value=element)
            last_row += 1

        print(f"Les éléments {elements_manquants} ont été ajoutés à la fin de la colonne A.")
    else:
        print("Aucun élément manquant trouvé.")

# Demander le chemin du fichier Excel à l'utilisateur
excel_file_path = input("Veuillez entrer le chemin du fichier Excel : ")

# Charger le fichier Excel
workbook = openpyxl.load_workbook(excel_file_path)

# Trouver le sheet avec le plus grand nombre de lignes
selected_sheet = find_sheet_with_most_rows(workbook)

if selected_sheet:
    sheet = workbook[selected_sheet]

    # Récupérer toutes les valeurs de la colonne A dans une liste
    column_A = [cell.value for cell in sheet['A']]

    # Liste à comparer
    list_to_compare = ["Test"]  # Remplace cela par ta propre liste

    # Comparer les listes et ajouter les éléments manquants
    ajouter_elements_manquants(workbook, selected_sheet, column_A, list_to_compare)

    # Sauvegarder les modifications dans le fichier Excel
    workbook.save(excel_file_path)
    print("Les modifications ont été sauvegardées.")
else:
    print("Aucun sheet trouvé.")

# Fermer le fichier Excel
workbook.close()
