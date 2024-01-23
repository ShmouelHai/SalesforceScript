import openpyxl
import xml.etree.ElementTree as ET
import pandas as pd
import os

def find_sheet_with_most_rows(workbook):
    sheets = workbook.sheetnames
    sheet_with_most_rows = None
    max_rows = 0

    for sheet_name in sheets:
        sheet = workbook[sheet_name]
        num_rows = sheet.max_row

        if num_rows > max_rows:
            max_rows = num_rows
            sheet_with_most_rows = sheet_name

    return sheet_with_most_rows

def ajouter_elements_manquants(workbook, selected_sheet, column_A, list_to_compare, listeDeListes):
    sheet = workbook[selected_sheet]


    # Vérifier les éléments manquants
    elements_manquants = [element for element in list_to_compare if element not in column_A]

    if elements_manquants:
        # Ajouter les éléments manquants à la fin de la colonne A
        last_row = sheet.max_row + 1
        for element in elements_manquants:
            for details in listeDeListes:
                if details[0] == element:
                    sheet.cell(row=last_row, column=1, value=details[0])
                    sheet.cell(row=last_row, column=2, value=details[1])
                    sheet.cell(row=last_row, column=3, value=details[2])
                    sheet.cell(row=last_row, column=4, value="added")
                    last_row += 1

        print(f"Les éléments {elements_manquants} ont été ajoutés à la fin de la colonne A.")
    else:
        print("Aucun élément manquant trouvé.")


#******************************* MAIN ***********************************

# Demander le chemin du fichier Excel à l'utilisateur
excel_file_path = input("Veuillez entrer le chemin du fichier Excel : ")

# Charger le fichier Excel
workbook = openpyxl.load_workbook(excel_file_path)

# Trouver le sheet avec le plus grand nombre de lignes
selected_sheet = find_sheet_with_most_rows(workbook)



if selected_sheet:
    sheet = workbook[selected_sheet]

    # Récupérer toutes les valeurs de la colonne A dans une liste
    column_A = [cell.value for cell in sheet['A']]

    # Liste à comparer
    xml_file_path = input("Entrez le chemin du fichier XML : ")
    # Parsing XML
    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    membersToRemove = ["et4ae5", "due__", "Google", "OSF_", "twilio", "odigo", "Didenjoy", "IndividualEmailResult", 
                       "ServiceTerritory", "ServiceResource", "SocialPost", "ServiceAppointment", "ResourceAbsence"
                       ,"WorkType", "SiqUserBlacklist", "SocialPost", "StreamActivity", "TableauHost", "UserEmailPreferred",
                       "VideoCall", "VoiceCall", "Waitlist"]
    membersList = []
    finalGrandList = []
    finalGrandListMembers = []
    statusMembersList = []

    df = pd.DataFrame()

    # Chaque type est défini par un name
    for types in root:
        nameTypeList = ''
        membersList = []
        for elem in types:
            flag = 0
            elem_name = elem.tag.split('}')[-1]
            if elem_name == 'members':
                membersList.append(elem.text)
                for indesirables in membersToRemove:
                    if indesirables in elem.text:
                        flag = 1 
                statusMembersList.append("Expected OK" if flag == 0 else "Not supported in SFOA")
            if elem_name == 'name':
                nameTypeList = elem.text

        if nameTypeList == 'EmailTemplate':
            statusMembersList[-len(membersList):] = ["Data Scope"] * len(membersList)

        if nameTypeList == 'Report':
            statusMembersList[-len(membersList):] = ["Scope To Define"] * len(membersList)

        if nameTypeList == 'Dashboard':
            statusMembersList[-len(membersList):] = ["Scope To Define"] * len(membersList)

        listeDeListes = []

        # Construction de finalGrandList après que statusMembersList a été construit 
        for member in membersList:
            finalGrandList.append(member)
            finalGrandListMembers.append(nameTypeList)

    for i in range(len(finalGrandList)):
        detailsMembre = [finalGrandList[i], finalGrandListMembers[i], statusMembersList[i]]
        # Ajout de la sous-liste à la liste principale
        listeDeListes.append(detailsMembre)

    # Création d'un DataFrame Pandas
    print("finalGrandList: " + str(len(finalGrandList)))
    print("finalGrandListMembers: " + str(len(finalGrandListMembers)))
    print("statusMembersList: " + str(len(statusMembersList)))
    print("listeDeListes: " + str(len(listeDeListes)))

    # df = pd.DataFrame({"Members": finalGrandList, "Type": finalGrandListMembers  , "SFOA Status": statusMembersList}) 

    list_to_compare = finalGrandList  # Remplace cela par ta propre liste

    # Comparer les listes et ajouter les éléments manquants
    ajouter_elements_manquants(workbook, selected_sheet, column_A, list_to_compare, listeDeListes)

    # Sauvegarder les modifications dans le fichier Excel
    workbook.save(excel_file_path)
    print("Les modifications ont été sauvegardées.")
else:
    print("Aucun sheet trouvé.")

# Fermer le fichier Excel
workbook.close()
